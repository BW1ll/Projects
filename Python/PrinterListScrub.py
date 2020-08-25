#!/usr/bin/env python3
# brian@lunaWHN.com
# 12/17/19

# for the spread sheets
import os
import openpyxl
from openpyxl import Workbook as wb


# gloval - lits other directorys as needed
workingDir = 'F:\\Python\\DataToPaperCut\\' # location of the files

# this is where the batch file must be run from on the PaperCut Server
executionDir = 'C:\\Program Files\\PaperCut MF\\server\\bin\\win\\'  ## PaperCut Server

# excel files to work with
aList = 'Asset_List.xlsx'
pList = 'printer_list.xlsx'

# global -
print('Opening workbook...')
wbaList = openpyxl.load_workbook(f'{workingDir}{aList}')
wbpList = openpyxl.load_workbook(f'{workingDir}{pList}')
ws = wbpList.active
clean = ['net://', 'Local://']

def printerListCleanup():
    ''' cleans up the inital data sheet. Removes 'net://, and Local://' from the thrid cell and deletes 
    the others that aren't needed from the report '''

    print('Cleaning datasheet...')
    for i in range(3, len(ws['C'])+1):
        cell = ws[f'C{i}']
        if clean in cell.value:  # clean device column
            temp = cell.value
            fixed = temp.replace(clean, '')
            cell.value = fixed

    # remove columns
    ws.delete_cols(14, 6) # starts at 14th column and delets 6
    ws.delete_cols(6, 7)  # starts at 6th column and delets 7
    wbpList.save(f'{workingDir}scrubedPrinterList.xlsx')





# runs main functions
if __name__ == "__main__":
    print('\nMake sure you have the correct directories listed in the script.')
    print('And you have the .xlsx version of the device list from Papercut in the working dir.')
    input('\nPress ENTER to continue...')
    printerListCleanup()  # clean the inital data sheet
    # deviceLoop()  # start the loop
    print('\nDone, son.\n')