#!/usr/bin/env python3
'''
output lines that are needed
server-command set-printer-property <ServerNam> "printer_name" custom-feild-6 'Cost Center'
server-command set-printer-property <ServerNam> "printer_name" custom-feild-4 'Serical Number'
server-command set-printer-property <ServerNam> "printer_name" custom-feild-3 'Equipment ID Number'
server-command set-printer-groups <ServerNam> "printer_name" 'Printer Groups'
needs to pause afte each entrie and wait for previous command to complete.
'''

import openpyxl as op
from openpyxl import Workbook

workingDir = 'F:\\Python\\DataToPaperCut\\'
spList = (f'{workingDir}scrubedPrinterList1.xlsx')
fileExel = (f'{workingDir}output.xlsx')

wb = op.load_workbook(spList)
ws = wb.active

wb2 = op.load_workbook(fileExel)
ws2 = wb2.active

oputFile = (f'{workingDir}batch.txt')

# df_initial = pd.read_excel(spList)

#  print(df_initial.columns)

# prinName = spList.
def cleanData():
 
    print('Cleaning datasheet...')
    for i in range(2, len(ws['F'])+1):
        cell = ws[f'F{i}']
        
        if '|' in cell.value:  # clean device column
            print('Fournd 1')
            temp = cell.value
            print(temp)
            fixed = temp.replace('|', 'TEST')
            print(fixed)
            cell = fixed
            print(cell)

    wb.save(ws2)

    

def outPut():
    fielObject = open(oputFile, 'a')
    for i in range(3, len(ws['J'])+1):
        cellB = ws2[f'B{i}'].value
        cellE = ws2[f'E{i}'].value
        cellF = ws2[f'F{i}'].value
        cellG = ws2[f'G{i}'].value
        cellJ = ws2[f'J{i}'].value
        # adds the Cost Center to the command
        fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-6 \"{cellJ}\"\n')
        fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-4 \"{cellE}\"\n')
        if cellF is not None:
            # adds the printer groups to the command
            fielObject.write(f'server-command set-printer-groups <ServerNam> {cellB} \"{cellF}\"\n')
            if cellG is not None: 
                # adds the Equipment ID number to the command
                fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-3 \"{cellG}\"\n')
                
                
                fielObject.close

        

if __name__ == "__main__":
    cleanData()
    
    #outPut()