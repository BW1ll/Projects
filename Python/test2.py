#!/usr/bin/env python3
'''
output lines that are needed:
server-command set-printer-property <ServerNam> "printer_name" custom-feild-6 'Cost Center'
server-command set-printer-property <ServerNam> "printer_name" custom-feild-4 'Serical Number'
server-command set-printer-property <ServerNam> "printer_name" custom-feild-3 'Equipment ID Number'
server-command set-printer-groups <ServerNam> "printer_name" 'Printer Groups'
need VBScript or Batch script to exicut the commands
needs to pause afte each entrie and wait for previous command to complete.
'''

import openpyxl as op
from openpyxl import Workbook

workingDir = 'F:\\Python\\DataToPaperCut\\'  # set the working file directory
spList = (f'{workingDir}PaperCut upload data.xlsx')  #set the working Excel sheet

wb = op.load_workbook(spList)  # loads in the excel file
wb.active = 1 # sets the active work sheet in the workbook
ws = wb.active

oputFile = (f'{workingDir}batch.txt')


def cleanData():  # use this function to clean up cell data if needed.
 
    print('Cleaning datasheet...')
    for i in range(2, len(ws['F'])+1):
        cell = ws[f'F{i}']
        
        if 'admin' in cell.value:  # clean device column
            print(f'Fournd {i}')
            temp = cell.value
            print(temp)
            fixed = temp.replace('admin', 'TEST')
            print(fixed)
            cell = fixed
            print(cell)


    wb.save(f'{workingDir}output.xlsx')

def outPut():
    fielObject = open(oputFile, 'a')
    for i in range(2, len(ws['J'])+1):
        cellB = ws[f'B{i}'].value
        cellE = ws[f'E{i}'].value
        cellF = ws[f'F{i}'].value
        cellG = ws[f'G{i}'].value
        cellJ = ws[f'J{i}'].value
        # adds the Cost Center to the command
        #fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-6 \"{cellJ}\"\n')
        #fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-4 \"{cellE}\"\n')
        if cellF is not None:
            # adds the printer groups to the command
            fielObject.write(f'server-command set-printer-groups <ServerNam> {cellB} \"{cellF}\"\n')
        #    if cellG is not None: 
        #        # adds the Equipment ID number to the command
        #        fielObject.write(f'server-command set-printer-property <ServerNam> {cellB} custom-field-3 \"{cellG}\"\n')
                
                
            fielObject.close

        

if __name__ == "__main__":
#    cleanData()
    outPut()