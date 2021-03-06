#!/usr/bin/env python3
# brian@lunaWHN.com

import openpyxl as op
from openpyxl import Workbook 

names = []
working_dir = 'C:\\Users\\BriWillis\\Downloads\\'
out_put_dir = 'C:\\projects\\MainCode\\Python\\DataToPaperCut\\PaperCutBatch\\'
file_excel = (f'{working_dir}user_list (1).xlsx')

wb = op.load_workbook(file_excel)
ws = wb.active
oputFile = (f'{out_put_dir}delete_names.txt')
   
print('Building Commands')
for value in range(1, len(ws['A'])+1):
    cell = ws[f'A{value}']
    names.append(cell.value)

print(names)
file_object = open(oputFile, 'w')

for name in names:
    file_object.write(f'server-command delete-existing-user {name}\n')
    file_object.close
